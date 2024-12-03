from django.shortcuts import render
from django.shortcuts import get_object_or_404, redirect
from voyage.forms import VoyageForm
from voyage.models import *
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib.auth.forms import AuthenticationForm
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import datetime, timedelta

@login_required(login_url='login')
def index(request):
    Voyages = Voyage.objects.all()

    context = {
        'Voyages': Voyages
    }
    return render(request, 'accueil.html', context)

@login_required(login_url='login')
def ajout_voyage(request):
    if request.method == 'POST':
        form = VoyageForm(request.POST)
        if form.is_valid():
            form.save()
            voyage = form.save(commit=False)
            voyage.user = request.user
            voyage.save()
            user = request.user
            if user.groups.filter(name='Employe').exists():
             return redirect('home')
            if user.groups.filter(name='Admin').exists():
                return redirect('home1')
    else:
        form = VoyageForm()
    return render(request, 'ajouter_voyage.html', {'form': form})

from django.contrib import messages

@login_required(login_url='login')
def supprimer_voyage(request, pk):
    voyage = get_object_or_404(Voyage, pk=pk, user=request.user)
    if request.user==voyage.user:
        voyage.delete()
        messages.success(request, "Voyage supprimé avec succès.")
        user=request.user
        if user.groups.filter(name='Employe').exists():
            return redirect('home')
        if user.groups.filter(name='Admin').exists():
            return redirect('home1')
    if request.user != voyage.user:
        return HttpResponse("Vous n'êtes pas autorisé à modifié ce voyage.")


@login_required(login_url='login')
def modifier_voyage(request, pk):
    voyage = Voyage.objects.get(pk=pk)
    if request.user==voyage.user:
        if request.method == 'POST':
            form = VoyageForm(request.POST, instance=voyage)
            user = request.user
            if form.is_valid():
                voyage = form.save(commit=False)
                voyage.user = request.user
                voyage.save()
                if user.groups.filter(name='Employe').exists():
                    return redirect('home')
                if user.groups.filter(name='Admin').exists():
                    return redirect('home1')
        else:
            form = VoyageForm(instance=voyage)
    if request.user!=voyage.user :
        return HttpResponse("Vous n'êtes pas autorisé à modifié ce voyage.")

    return render(request, 'modifier_voyage.html', {'form': form})



@login_required(login_url='login')
def export_voyages_to_excel(request):
    # Créer un fichier Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Voyages"

    # Ajouter les en-têtes de colonne
    headers = [
        "Chauffeur", "Numéro de plaque", "Nombre de places vides",
        "Ville de départ", "Ville d'arrivée", "Heure de départ",
        "Prix"
    ]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Ajouter les données
    voyages = Voyage.objects.all()
    for row_num, voyage in enumerate(voyages, 2):
        sheet.cell(row=row_num, column=1, value=voyage.NomChauffeur)
        sheet.cell(row=row_num, column=2, value=voyage.plaque)
        sheet.cell(row=row_num, column=3, value=voyage.NbrPlaceVide)
        sheet.cell(row=row_num, column=4, value=voyage.VilleDepart)
        sheet.cell(row=row_num, column=5, value=voyage.VilleArrivee)
        sheet.cell(row=row_num, column=6, value=str(voyage.heureDepart))
        sheet.cell(row=row_num, column=7, value=voyage.Prix)

    # Ajuster la largeur des colonnes
    for col_num, _ in enumerate(headers, 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = 20

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=liste des voyages .xlsx'
    workbook.save(response)
    return response

@login_required(login_url='login')
def chart(request):
    return render(request, 'voyageChart.html')

@login_required(login_url='login')
def employe(request):
    # Filtrer les voyages par utilisateur connecté
    voyages = Voyage.objects.filter(user=request.user)
    escales = Escale.objects.filter(user=request.user)

    # Définir la date limite pour les voyages récents
    days_ago = timezone.now() - timedelta(days=30)

    # Annoter les voyages pour ajouter un champ "recent" dans la requête
    voyages = voyages.annotate(
        recent=Case(
            When(date__gte=days_ago, then=Value(True)),  # Si la date est >= il y a 2 jours
            default=Value(False),  # Sinon False
            output_field=BooleanField(),
        )
    )
    # Filtrage par ville de départ et d'arrivée
    ville_depart = request.GET.get('ville_depart')
    ville_arrivee = request.GET.get('ville_arrivee')

    if ville_depart:
        voyages = voyages.filter(VilleDepart=ville_depart)
    if ville_arrivee:
        voyages = voyages.filter(VilleArrivee=ville_arrivee)

    # Tri
    sort_by = request.GET.get('sort_by', 'date')  # Tri par défaut par date
    if sort_by in ['date', 'Prix', 'NbrPlaceVide']:
        voyages = voyages.order_by(sort_by)

    # Pagination (10 voyages par page)
    paginator = Paginator(voyages, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'Voyages': page_obj,
        'ville_depart': ville_depart,
        'ville_arrivee': ville_arrivee,
        'sort_by': sort_by,
        'escales':escales
    }
    return render(request, 'index.html', context)

from django.db.models import BooleanField, Case, When, Value

@login_required(login_url='login')
def adminPage(request):
    # Filtrer les voyages par utilisateur connecté
    voyages = Voyage.objects.filter(user=request.user)
    escales = Escale.objects.filter(voyage__user=request.user)

    # Définir la date limite pour les voyages récents
    days_ago = timezone.now() - timedelta(days=2)

    # Annoter les voyages pour ajouter un champ "recent" dans la requête
    voyages = voyages.annotate(
        recent=Case(
            When(date__gte=days_ago, then=Value(True)),  # Si la date est >= il y a 2 jours
            default=Value(False),  # Sinon False
            output_field=BooleanField(),
        )
    )
    # Filtrage par ville de départ et d'arrivée
    ville_depart = request.GET.get('ville_depart')
    ville_arrivee = request.GET.get('ville_arrivee')

    if ville_depart:
        voyages = voyages.filter(VilleDepart=ville_depart)
    if ville_arrivee:
        voyages = voyages.filter(VilleArrivee=ville_arrivee)

    # Tri
    sort_by = request.GET.get('sort_by', 'date')  # Tri par défaut par date
    if sort_by in ['date', 'Prix', 'NbrPlaceVide']:
        voyages = voyages.order_by(sort_by)

    # Pagination (10 voyages par page)
    paginator = Paginator(voyages, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    today_minus_seven_days = timezone.now() - timedelta(days=7)

    context = {
        'Voyages': page_obj,
        'ville_depart': ville_depart,
        'ville_arrivee': ville_arrivee,
        'sort_by': sort_by,
        'escales':escales,
    }
    return render(request, 'accueil.html', context)

def login_view(request):
    if request.user.is_authenticated:
        if request.user.groups.filter(name='Employe').exists():
            login(request, request.user)
            return redirect('home')
        elif request.user.groups.filter(name='Admin').exists():
            login(request, request.user)
            return redirect('home1')
        else:
            return redirect('')
    else :
        if request.method == 'POST':
            form = AuthenticationForm(request, data=request.POST)
            if form.is_valid():
                # Authentifier l'utilisateur
                username = form.cleaned_data.get('username')
                password = form.cleaned_data.get('password')
                user = authenticate(username=username, password=password)
                if user is not None :
                    if user.groups.filter(name='Employe').exists():
                        login(request, user)
                        return redirect('home')
                    elif user.groups.filter(name='Admin').exists():
                        login(request, user)
                        return redirect('home1')
                    else:
                        return redirect('')
                else:
                    return HttpResponse("Utilisateur non trouvé.", status=401)
            else:
                messages.error(request, "Nom d'utilisateur ou mot de passe incorrect, Veuillez réessayer.")
                return redirect('login')
        else:
            form = AuthenticationForm()

        return render(request,'account/login.html',{'form': form})

def logout_view(request):
    # Déconnexion de l'utilisateur
    logout(request)
    # Redirection vers la page de connexion ou autre page
    return redirect('login')


def home0(request):
    if request.user.groups.filter(name='Employe').exists():
        login(request, request.user)
        return redirect('home')
    elif request.user.groups.filter(name='Admin').exists():
        login(request, request.user)
        return redirect('home1')
    else:
        return redirect('')